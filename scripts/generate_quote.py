import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

data = [
["1","云安全中心-标准版","包含日志分析量为每月50G；云安全中心系统主要包含安全态势、资产中心、风险管理、威胁管理、分析中心、告警管理、编排响应、报表中心、集成配置以及数据源监控等功能。",1600,19200,16320,7344,8976,"包年订购折扣：针对一次性包年付费服务，享受1年及以上8.5折优惠。","0.85",""],
["2","云安全中心-日志分析量","默认需购买500G日志分析量，额外购买步长为50G",225,2700,2295,1033,1262,"包年订购折扣：针对一次性包年付费服务，享受1年及以上8.5折优惠。","0.85",""],
["3","服务安全卫士-旗舰版","按照需防护的主机个数购买，提供安全概览、资产管理、入侵检测、漏洞扫描、基线管理功能。",180,2160,1836,826,1010,"包年订购折扣：针对一次性包年付费服务，1年85折，2年7折，3、4、5年5折。","0.85",""],
["4","Web应用防火墙-独享版-单机版","支持0-1Gbps的业务防护，默认业务请求峰值3000QPS（200Mbps），默认防护域名（不区分主域名、子域名）/IP（公网IP、私网IP）个数100个/实例，防护端口数最大支持65535个（常用端口除外）",2837,34044,34044,15320,18724,"-","1.0",""],
["5","下一代防火墙-标准版","支持1Gbps公网流量处理能力峰值（每秒数据包处理能力：200000 PPS）；提供云上互联网边界和VPC边界的防护，包括：实时入侵检测与防御、全局统一访问控制、全流量分析可视化、日志审计与溯源分析。",1879,22548,22548,10147,12401,"-","1.0",""],
["6","数据库审计-4资产","支持4数据库实例；通过Agent抓包方式旁路部署，提供数据库审计、SQL注入攻击检测、风险操作识别等功能，保障云上数据库的安全。",3333,39996,33317,14993,18324,"包年订购折扣：针对一次性包年付费服务，享受1年及以上8.33折优惠。","0.833",""],
["7","日志审计-10资产","支持10个日志源（每个日志源对应为一个IP）；通过主被动结合的方式，实时不间断地采集用户网络中各种不同厂商的安全设备、网络设备、主机、操作系统、以及各种应用系统产生的海量日志信息，并出具丰富的报表报告。",1107,13284,11291,5081,6210,"包年订购折扣：针对一次性包年付费服务，1年85折，2年75折，3、4、5年65折。","0.85",""],
["8","堡垒机-10资产","支持10资产管理；4A统一安全管控平台，为企业提供集单点登录、统一资产管理、多终端访问协议、文件传输、会话协同等功能于一体的运维管理服务。",1020,12240,10404,4682,5722,"包年订购折扣：针对一次性包年付费服务，1年85折，2年7折，3、4、5年5折。","0.85",""],
["9","漏洞扫描-10资产","支持10个IP地址；能够对Web应用的资产进行识别分类以及对Web应用进行深度弱点探测。",937,11244,11244,5060,6184,"","1.0",""]
]

cols = ["序号","产品规格","详细规格说明","标准价(元/月)","标准价(元/年)","折扣价(包年)","4.5折结算价","5.5折结算价","包年优惠","1年折扣率","备注"]

df = pd.DataFrame(data, columns=cols)
# total row
total = {
    "序号":"合计",
    "产品规格":"",
    "详细规格说明":"",
    "标准价(元/月)":sum([row[3] for row in data]),
    "标准价(元/年)":sum([row[4] for row in data]),
    "折扣价(包年)":sum([row[5] for row in data]),
    "4.5折结算价":sum([row[6] for row in data]),
    "5.5折结算价":sum([row[7] for row in data]),
    "包年优惠":"",
    "1年折扣率":"",
    "备注":""
}

df = pd.concat([df, pd.DataFrame([total])], ignore_index=True)

output_path = r"C:\Users\lpcis\WorkBuddy\Ctyun-quote\quote\等保专区-三级等保基础版报价.xlsx"

df.to_excel(output_path, index=False, sheet_name="报价")

# Formatting
wb = load_workbook(output_path)
ws = wb.active
header_fill = PatternFill(start_color="3B5998", end_color="3B5998", fill_type="solid")
header_font = Font(name="微软雅黑", size=12, bold=True, color="FFFFFFFF")
header_alignment = Alignment(horizontal="center", vertical="center")
thin = Side(border_style="thin", color="000000")
border = Border(left=thin, right=thin, top=thin, bottom=thin)
for col in range(1, ws.max_column+1):
    cell = ws.cell(row=1, column=col)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = header_alignment
    cell.border = border
    ws.column_dimensions[get_column_letter(col)].width = 15
for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
    for cell in row:
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        cell.border = border
ws.freeze_panes = ws["A2"]
# add note row
note = "服务商：LX-2；需要绑定弹性IP进行管理，成功开通后请勿解绑该IP；需要单独创建一个安全子网（至少保留9个以上可用IP地址）用于部署云等保专区，不能与业务主机所在子网相同，否则可能会存在无法交付的问题。"
ws.append([note])
note_row = ws.max_row
ws.merge_cells(start_row=note_row, start_column=1, end_row=note_row, end_column=ws.max_column)
ws.cell(row=note_row, column=1).alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
wb.save(output_path)
