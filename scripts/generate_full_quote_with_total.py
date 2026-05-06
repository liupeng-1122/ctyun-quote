import pandas as pd
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

# Data rows for 等保专区 三级等保基础版
rows = [
    {"序号":1, "产品规格":"安全中心（云安全中心‑标准版）", "详细规格说明":"包含日志分析量每月 50 GB，提供安全态势感知、资产中心、风险管理、威胁管理、分析中心、告警管理、编排响应、报表中心、集成配置、数据源监控等功能。", "标准价（1 月）":1600, "标准价（1 年）":19200, "折扣价（包年）":16320, "结算价 4.5折":7344, "结算价 5.5折":8976, "备注":"包年1年85折，2年7折，3年5折"},
    {"序号":2, "产品规格":"云安全中心‑日志分析量", "详细规格说明":"默认购买 500 GB 日志分析量（可按 50 GB 步长增购）。", "标准价（1 月）":225, "标准价（1 年）":2700, "折扣价（包年）":2295, "结算价 4.5折":1033, "结算价 5.5折":1262, "备注":"包年1年85折，2年7折，3年5折"},
    {"序号":3, "产品规格":"服务安全卫士‑旗舰版", "详细规格说明":"全方位服务器安全防护，资产清点、漏洞扫描、入侵检测、基线检查、弱口令检测、病毒查杀、文件完整性保护、勒索诱饵、蜜罐防护、端口蜜罐、日志告警、自动化处置，支持 Windows、Linux、容器。", "标准价（1 月）":180, "标准价（1 年）":2160, "折扣价（包年）":1836, "结算价 4.5折":826, "结算价 5.5折":1010, "备注":"包年1年85折，2年7折，3年5折"},
    {"序号":4, "产品规格":"Web应用防火墙‑独享版‑单机版", "详细规格说明":"支持 0‑1 Gbps 防护，默认 100 域名/IP，最大 65,535 端口，防御 Web 攻击、CC、数据泄露、SQL 注入、跨站脚本、文件上传风险，自动化规则更新、日志审计。", "标准价（1 月）":2837, "标准价（1 年）":34044, "折扣价（包年）":34044, "结算价 4.5折":15320, "结算价 5.5折":18724, "备注":"无包年折扣（按原价计）"},
    {"序号":5, "产品规格":"下一代防火墙‑标准版", "详细规格说明":"1 Gbps 公网流量处理能力，200,000 PPS，实时入侵检测与防御、全局统一访问控制、全流量可视化、日志审计与溯源，支持 VPC 边界防护。", "标准价（1 月）":1879, "标准价（1 年）":22548, "折扣价（包年）":22548, "结算价 4.5折":10147, "结算价 5.5折":12401, "备注":"无包年折扣（按原价计）"},
    {"序号":6, "产品规格":"数据库审计‑4资产", "详细规格说明":"支持 4 个数据库实例（MySQL、PostgreSQL、SQL Server、Oracle），采用 Agent 被动/主动抓包，提供审计、SQL 注入检测、风险操作识别、审计日志存储、查询、报表。", "标准价（1 月）":3333, "标准价（1 年）":39996, "折扣价（包年）":33317, "结算价 4.5折":14993, "结算价 5.5折":18324, "备注":"包年1年8.33折，2年7折，3年5折"},
    {"序号":7, "产品规格":"日志审计‑10资产", "详细规格说明":"支持 10 条日志源（每条对应一个 IP），实时采集云上网络、主机、应用日志，集中存储、索引、审计、告警、响应、报表。", "标准价（1 月）":1107, "标准价（1 年）":13284, "折扣价（包年）":11291, "结算价 4.5折":5081, "结算价 5.5折":6210, "备注":"包年1年85折，2年75折，3年65折"},
    {"序号":8, "产品规格":"堡垒机‑10资产", "详细规格说明":"支持 10 资产管理，统一安全平台，单点登录、统一资产管理、多协议访问（SSH、RDP、VNC、SFTP、FTP），会话协同、文件传输、命令控制、工单审批、细粒度权限、审计录像。", "标准价（1 月）":1020, "标准价（1 年）":12240, "折扣价（包年）":10404, "结算价 4.5折":4682, "结算价 5.5折":5722, "备注":"包年1年85折，2年7折，3年5折"},
    {"序号":9, "产品规格":"漏洞扫描‑10资产", "详细规格说明":"支持 10 个 IP，深度 Web 应用弱点探测，覆盖常见 CMS、操作系统等漏洞，提供端口服务识别、风险报告。", "标准价（1 月）":937, "标准价（1 年）":11244, "折扣价（包年）":11244, "结算价 4.5折":5060, "结算价 5.5折":6184, "备注":"无包年折扣（按原价计）"}
]

# Compute totals
import numpy as np

df = pd.DataFrame(rows)
# Sum numeric columns
numeric_cols = ["标准价（1 月）", "标准价（1 年）", "折扣价（包年）", "结算价 4.5折", "结算价 5.5折"]
totals = df[numeric_cols].sum().to_dict()
# Append total row
total_row = {"序号":"合计", "产品规格":"", "详细规格说明":"", "标准价（1 月）":totals["标准价（1 月）"], "标准价（1 年）":totals["标准价（1 年）"], "折扣价（包年）":totals["折扣价（包年）"], "结算价 4.5折":totals["结算价 4.5折"], "结算价 5.5折":totals["结算价 5.5折"], "备注":""}
df = pd.concat([df, pd.DataFrame([total_row])], ignore_index=True)

output_path = r"c:/Users/lpcis/WorkBuddy/Ctyun-quote/quote/等保专区三级等保基础版报价_v2.xlsx"
with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
    df.to_excel(writer, index=False, sheet_name='报价')
    wb = writer.book
    ws = wb['报价']
    # Header style
    header_font = Font(name='微软雅黑', size=12, bold=True, color='FFFFFFFF')
    header_fill = PatternFill(fill_type='solid', fgColor='3B5998')
    alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin = Side(border_style='thin', color='000000')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = alignment
        cell.border = border
    # Column widths
    ws.column_dimensions['A'].width = 6
    for col in ['B','C','D','E','F','G','H','I','J']:
        ws.column_dimensions[col].width = 15
    ws.freeze_panes = ws['A2']
    # Body style
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=10):
        for cell in row:
            cell.font = Font(name='微软雅黑', size=11)
            cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
            cell.border = border
    # Highlight discount columns G,H,I
    for r in range(2, ws.max_row+1):
        ws[f'G{r}'].fill = PatternFill(fill_type='solid', fgColor='FFF2CC')
        ws[f'H{r}'].fill = PatternFill(fill_type='solid', fgColor='FFF2CC')
        ws[f'I{r}'].fill = PatternFill(fill_type='solid', fgColor='FFF2CC')
print('Excel generated with totals')
