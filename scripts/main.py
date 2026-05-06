# -*- coding: utf-8 -*-
"""
天翼云等保专区 - 报价表生成器
支持合营池云主机绑定功能
"""
import os
import sys
import argparse
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# 导入报价模块
from scripts.quote_generator import (
    get_all_products, get_all_hosts, get_product_by_id, 
    get_basic_edition_products, calculate_price, generate_quote_data, get_totals
)

# ========== Excel 样式定义 ==========
HEADER_FONT = Font(name='微软雅黑', size=12, bold=True, color='FFFFFF')
HEADER_FILL = PatternFill(start_color='3B59B8', end_color='3B59B8', fill_type='solid')
HEADER_ALIGN = Alignment(horizontal='center', vertical='center', wrap_text=True)

BODY_FONT = Font(name='微软雅黑', size=10.5)
BODY_ALIGN_LEFT = Alignment(horizontal='left', vertical='center', wrap_text=True)
BODY_ALIGN_CENTER = Alignment(horizontal='center', vertical='center', wrap_text=True)
BODY_ALIGN_RIGHT = Alignment(horizontal='right', vertical='center')

TOTAL_FILL = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')

THIN = Side(style='thin')
THIN_BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

# 注意事项
NOTES = """注意事项：
1. 服务商：LX-2；
2. 需要绑定弹性IP进行管理，成功开通后请勿解绑该IP；
3. 需要单独创建一个安全子网（至少保留9个以上可用IP地址）用于部署云等保专区，不能与业务主机所在子网相同，否则可能会存在无法交付的问题。"""

script_dir = os.path.dirname(os.path.abspath(__file__))


def create_quote_excel(data, totals, output_path, title="天翼云等保专区安全产品报价表", include_notes=True, years=1):
    """生成Excel报价表"""
    wb = Workbook()
    ws = wb.active
    ws.title = "报价表"
    
    # 列宽设置
    col_widths = {
        'A': 8,   # 序号
        'B': 20,  # 产品规格
        'C': 50,  # 详细规格说明
        'D': 8,   # 数量
        'E': 14,  # 标准价(1个月)
        'F': 14,  # 标准价(1年)
        'G': 14,  # 折扣价(1年)
        'H': 14,  # 4.5折结算价
        'I': 14,  # 5.5折结算价
        'J': 18,  # 包年优惠
        'K': 15,  # 备注
    }
    
    for col, width in col_widths.items():
        ws.column_dimensions[col].width = width
    
    # 表头（动态标注年数）
    headers = ['序号', '产品规格', '详细规格说明', '数量', '标准价\n(1个月)',
               f'标准价\n({years}年)',
               f'折扣价\n({years}年)',
               f'4.5折结算价\n({years}年)',
               f'5.5折结算价\n({years}年)',
               '包年优惠', '备注']
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN
        cell.border = THIN_BORDER
    
    ws.row_dimensions[1].height = 35
    
    # 数据行：按序号(seq)升序排列
    sorted_data = sorted(data, key=lambda x: x.get('seq', 0))
    for row_idx, item in enumerate(sorted_data, 2):
        y = item.get('yearly', 0)
        d = item.get('discounted', 0)
        p45 = item.get('price_45', 0)
        p55 = item.get('price_55', 0)
        row_data = [
            None,  # 序号列暂留，由下方单独写入公式
            item['product'],
            item['spec'],
            item['qty'],
            item['monthly'],
            y * years,
            d * years,
            p45 * years,
            p55 * years,
            item['discount_desc'],
            item['remark']
        ]
        
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = BODY_FONT
            cell.border = THIN_BORDER
            
            # 格式设置
            if col_idx == 1:  # 序号：写入公式 =ROW()-1
                cell.value = '=ROW()-1'
                cell.alignment = BODY_ALIGN_CENTER
            elif col_idx in [2, 3, 10, 11]:  # 文本
                cell.alignment = BODY_ALIGN_LEFT
            elif col_idx in [4]:  # 数量
                cell.alignment = BODY_ALIGN_CENTER
            else:  # 金额
                cell.alignment = BODY_ALIGN_RIGHT
                if isinstance(value, (int, float)):
                    cell.number_format = '#,##0.00'
            
            # 云主机行底色
            if item.get('is_host'):
                cell.fill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
        
        ws.row_dimensions[row_idx].height = 30
    
    # 合计行
    total_row = len(data) + 2
    total_data = ['合计', '', '', '', totals['monthly'], totals['yearly'] * years, 
                  totals['discounted'] * years, totals['price_45'] * years, totals['price_55'] * years, '', '']
    
    for col_idx, value in enumerate(total_data, 1):
        cell = ws.cell(row=total_row, column=col_idx, value=value)
        cell.font = Font(name='微软雅黑', size=11, bold=True)
        cell.fill = TOTAL_FILL
        cell.border = THIN_BORDER
        
        if col_idx == 1:
            cell.alignment = BODY_ALIGN_CENTER
        elif col_idx in [5, 6, 7, 8, 9]:
            cell.alignment = BODY_ALIGN_RIGHT
            if isinstance(value, (int, float)):
                cell.number_format = '#,##0.00'
        else:
            cell.alignment = BODY_ALIGN_LEFT
    
    # 注意事项
    if include_notes:
        notes_row = total_row + 2
        ws.cell(row=notes_row, column=1, value=NOTES).alignment = BODY_ALIGN_LEFT
        ws.row_dimensions[notes_row].height = 60
        # 合并单元格
        ws.merge_cells(start_row=notes_row, start_column=1, end_row=notes_row, end_column=11)
    
    # 冻结首行
    ws.freeze_panes = 'A2'
    
    wb.save(output_path)
    print(f"OK: {output_path}")


def list_all_products():
    """列出所有产品"""
    products = get_all_products()
    hosts = get_all_hosts()
    
    print("\n========== 安全产品 ==========")
    categories = {}
    for p in products:
        cat = p['category']
        if cat not in categories:
            categories[cat] = []
        categories[cat].append(p)
    
    for cat, prods in categories.items():
        print(f"\n【{cat}】")
        for p in prods:
            print(f"  [{p['id']}] {p['name']} - ¥{p['yearly_price']}/年")
    
    print("\n========== 云主机 ==========")
    for h in hosts:
        print(f"  [{h['id']}] {h['name']} - ¥{h['yearly_price']}/年")


def interactive_select():
    """交互式选择产品"""
    products = get_all_products()
    
    print("\n========== 选择产品 ==========")
    categories = {}
    for p in products:
        cat = p['category']
        if cat not in categories:
            categories[cat] = []
        categories[cat].append(p)
    
    for cat, prods in categories.items():
        print(f"\n【{cat}】")
        for p in prods:
            print(f"  [{p['id']}] {p['name']} - ¥{p['yearly_price']}/年")
    
    print("\n请输入产品ID（多个用逗号分隔，如: 1,4,14）: ", end="")
    ids_input = input().strip()
    
    selected = []
    for sid in ids_input.split(','):
        sid = sid.strip()
        if sid.isdigit():
            p = get_product_by_id(int(sid))
            if p:
                selected.append(p)
    
    # 设置数量
    quantities = {}
    if selected:
        print("设置数量（直接回车默认为1）:")
        for p in selected:
            print(f"  {p['name']}: ", end="")
            qty = input().strip()
            quantities[p['id']] = int(qty) if qty.isdigit() else 1
    
    return selected, quantities


def main():
    parser = argparse.ArgumentParser(description='天翼云等保专区报价表生成器')
    parser.add_argument('--list', action='store_true', help='列出所有产品')
    parser.add_argument('--basic', action='store_true', help='生成三级等保基础版')
    parser.add_argument('--full', action='store_true', help='生成全产品报价表')
    parser.add_argument('--output', '-o', type=str, help='输出文件路径')
    parser.add_argument('--products', type=str, help='指定产品ID（逗号分隔，如：1,3,5）')
    parser.add_argument('--qty', type=str, help='指定产品数量（逗号分隔，与--products配合）')
    parser.add_argument('--host', action='store_true', help='包含云主机（合营池模式）')
    parser.add_argument('--mgmt-host', action='store_true', help='包含管理节点云主机')
    parser.add_argument('--interactive', '-i', action='store_true', help='交互式选择产品')
    args = parser.parse_args()
    
    # 列表模式
    if args.list:
        list_all_products()
        return
    
    # 交互式模式
    if args.interactive:
        selected, qty_map = interactive_select()
        data = generate_quote_data(selected, qty_map, args.host, args.mgmt_host)
        totals = get_totals(data)
        
        quote_dir = os.path.join(os.path.dirname(script_dir), 'quote')
        os.makedirs(quote_dir, exist_ok=True)
        filename = f"天翼云等保专区安全产品报价表_自定义_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        output_path = os.path.join(quote_dir, filename)
        
        create_quote_excel(data, totals, output_path, "天翼云等保专区安全产品报价表（自定义）")
        print_summary(data, totals)
        return
    
    # 自定义产品模式
    if args.products:
        product_ids = [int(x.strip()) for x in args.products.split(',')]
        quantities = {}
        if args.qty:
            qtys = [int(x.strip()) for x in args.qty.split(',')]
            for i, pid in enumerate(product_ids):
                if i < len(qtys):
                    quantities[pid] = qtys[i]
        
        selected = [get_product_by_id(pid) for pid in product_ids if get_product_by_id(pid)]
        data = generate_quote_data(selected, quantities, args.host, args.mgmt_host)
        totals = get_totals(data)
        
        quote_dir = os.path.join(os.path.dirname(script_dir), 'quote')
        os.makedirs(quote_dir, exist_ok=True)
        filename = f"天翼云等保专区安全产品报价表_自定义_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        output_path = os.path.join(quote_dir, filename)
        
        create_quote_excel(data, totals, output_path, "天翼云等保专区安全产品报价表（自定义）")
        print_summary(data, totals)
        return
    
    # 基础版模式（默认）
    selected = get_basic_edition_products()
    quantities = {}
    if args.qty:
        qtys = [int(x.strip()) for x in args.qty.split(',')]
        for i, p in enumerate(selected):
            if i < len(qtys):
                quantities[p['id']] = qtys[i]
    
    data = generate_quote_data(selected, quantities, args.host, args.mgmt_host)
    totals = get_totals(data)
    
    quote_dir = os.path.join(os.path.dirname(script_dir), 'quote')
    os.makedirs(quote_dir, exist_ok=True)
    
    if args.output:
        output_path = args.output
    else:
        filename = f"天翼云等保专区安全产品报价表_基础版_{datetime.now().strftime('%Y%m%d')}.xlsx"
        output_path = os.path.join(quote_dir, filename)
    
    create_quote_excel(data, totals, output_path, "天翼云等保专区安全产品报价表（基础版）")
    print_summary(data, totals)


def print_summary(data, totals):
    """打印汇总"""
    print("\n========== 汇总 ==========")
    print(f"产品数量: {len(data)}")
    print(f"标准价(1年): ¥{totals['yearly']:,.2f}")
    print(f"折扣价(1年): ¥{totals['discounted']:,.2f}")
    print(f"4.5折结算: ¥{totals['price_45']:,.2f}")
    print(f"5.5折结算: ¥{totals['price_55']:,.2f}")


if __name__ == '__main__':
    main()
